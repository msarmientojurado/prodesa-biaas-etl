from openpyxl.workbook import Workbook
from openpyxl.styles import Font, colors, Color, Alignment, PatternFill, GradientFill, Border, Side
from openpyxl.styles import NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
import datetime

from tempfile import NamedTemporaryFile

from google.cloud import storage
from libraries.settings import BUCKET_NAME_DOWNLOAD_REPORT

def rpt_ar_deliveries_file(tbl_reporte_por_entregas_filter):
    
    tbl_reporte_por_entregas_excel=tbl_reporte_por_entregas_filter.reindex(columns=[
                                                                    'trpe_programacion',
                                                                    'trpe_tarea_entrega',
                                                                    'trpe_entrega_real',
                                                                    'trpe_entrega_programada'
                                                                    ])
    
    wb=Workbook()
    ws=wb.active

    chart_title = NamedStyle(name= 'title')
    chart_title.font=Font(bold=True, color='0000FF', size=14)
    chart_title.alignment=Alignment(horizontal='left', vertical='center')

    chart_date =NamedStyle(name='chart_date')
    chart_date.font=Font(bold=True, color='000000', size=12)
    chart_date.alignment=Alignment(horizontal='left', vertical='center')

    column_title =NamedStyle(name='column_title')
    column_title.font=Font(bold=True, color='000000', size=10)
    column_title.alignment=Alignment(horizontal='center', vertical='center', wrap_text=True)
    column_title.fill= PatternFill('solid',fgColor='D3BC5F')
    bd=Side(style='thick', color='000000')
    column_title.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    table_body =NamedStyle(name='table_body')
    table_body.font=Font(color='000000', size=10)
    table_body.alignment=Alignment(horizontal='left', vertical='center', wrap_text=False)
    bd=Side(style='thin', color='000000')
    table_body.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    table_body_centered =NamedStyle(name='table_body_centered')
    table_body_centered.font=Font(color='000000', size=10)
    table_body_centered.alignment=Alignment(horizontal='center', vertical='center', wrap_text=True)
    bd=Side(style='thin', color='000000')
    table_body_centered.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    table_body_red=NamedStyle(name='table_body_red')
    table_body_red.font=Font(color='000000', size=10)
    table_body_red.alignment=Alignment(horizontal='left', vertical='center', wrap_text=False)
    table_body_red.fill= PatternFill('solid',fgColor='FF0000')
    bd=Side(style='thin', color='000000')
    table_body_red.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    table_body_yellow=NamedStyle(name='table_body_yellow')
    table_body_yellow.font=Font(color='000000', size=10)
    table_body_yellow.alignment=Alignment(horizontal='left', vertical='center', wrap_text=False)
    table_body_yellow.fill= PatternFill('solid',fgColor='FFFF00')
    bd=Side(style='thin', color='000000')
    table_body_yellow.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    table_body_green=NamedStyle(name='table_body_green')
    table_body_green.font=Font(color='000000', size=10)
    table_body_green.alignment=Alignment(horizontal='left', vertical='center', wrap_text=False)
    table_body_green.fill= PatternFill('solid',fgColor='008000')
    bd=Side(style='thin', color='000000')
    table_body_green.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    table_body_text_red=NamedStyle(name='table_body_text_red')
    table_body_text_red.font=Font(color='FF0000', size=10)
    table_body_text_red.alignment=Alignment(horizontal='left', vertical='center', wrap_text=False)
    bd=Side(style='thin', color='000000')
    table_body_text_red.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    table_body_text_yellow=NamedStyle(name='table_body_text_yellow')
    table_body_text_yellow.font=Font(color='FFFF00', size=10)
    table_body_text_yellow.alignment=Alignment(horizontal='left', vertical='center', wrap_text=False)
    bd=Side(style='thin', color='000000')
    table_body_text_yellow.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    table_body_text_green=NamedStyle(name='table_body_text_green')
    table_body_text_green.font=Font(color='008000', size=10)
    table_body_text_green.alignment=Alignment(horizontal='left', vertical='center', wrap_text=False)
    bd=Side(style='thin', color='000000')
    table_body_text_green.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    bd_thick=Side(style='thick', color='000000')
    bd_thin=Side(style='thin', color='000000')

    table_body_total =NamedStyle(name='table_body_total')
    table_body_total.font=Font(color='000000', size=10)
    table_body_total.alignment=Alignment(horizontal='left', vertical='center', wrap_text=False)
    bd=Side(style='thin', color='000000')
    table_body_total.border = Border(left=bd_thin, top=bd_thin, right=bd_thin, bottom=bd_thick)

    ws['A1'].style = chart_title
    ws['A2'].style = chart_date
    ws['B2'].style = chart_date



    ws['A1']='Reporte por Entregas Proyecto '+(tbl_reporte_por_entregas_filter['trpe_proyecto'])[0]+' Etapa '+(tbl_reporte_por_entregas_filter['trpe_etapa'])[0]
    ws['A2']='Fecha Corte'
    ws['B2']='-'
    
#
    ws.merge_cells("A1:G1")
    ws.merge_cells("B2:C2")

    col_titles=tbl_reporte_por_entregas_excel['trpe_programacion'].unique()

    for c_idx, col_title in enumerate(col_titles, 2):
        cell=ws.cell(row=4, column=c_idx, value=col_title)
    
    
    tbl_reporte_por_entregas_excel=tbl_reporte_por_entregas_excel.sort_values(by=['trpe_tarea_entrega'],ascending=False)
    rows=dataframe_to_rows(tbl_reporte_por_entregas_excel, index=False)

    iteration=1
    programming = ""
    row_counter = 5
    first_iteration=True
    for r_idx, row in enumerate(rows,5):
        for c_idx, col in enumerate(row, 1):
            if first_iteration != True:
                if iteration == 1:
                    programming=col
                elif iteration == 2:
                    cell_bordered = ws.cell(row=row_counter, column=1, value= col)
                    cell_bordered.style=table_body_centered
                    cells_to_merge="A"+str(row_counter)+":A"+str(row_counter+1)
                    ws.merge_cells(cells_to_merge)
                elif iteration ==3:
                    col_counter=2
                    cell=ws.cell(row=4, column=col_counter)
                    while cell.value != None:
                        cell_bordered = ws.cell(row = row_counter, column= col_counter)
                        cell_bordered.style=table_body_centered
                        if cell.value == programming:
                            try:
                                info=col.strftime('%d-%m-%Y')
                            except:
                                info=""
                            ws.cell(row = row_counter, column= col_counter, value = info)
                        col_counter = col_counter + 1
                        cell=ws.cell(row=4, column=col_counter)
                elif iteration ==4:
                    col_counter=2
                    cell=ws.cell(row=4, column=col_counter)
                    while cell.value != None:
                        cell_bordered = ws.cell(row = row_counter+1, column= col_counter)
                        cell_bordered.style=table_body_centered
                        if cell.value == programming:
                            try:
                                info=col.strftime('%d-%m-%Y')
                            except:
                                info=""
                            ws.cell(row = row_counter +1, column= col_counter, value = info)
                        col_counter = col_counter + 1
                        cell=ws.cell(row=4, column=col_counter)
                
#
                iteration = iteration + 1
        iteration =  1
        row_counter = row_counter + 2
        first_iteration = False

    tmp = NamedTemporaryFile()
    wb.save(tmp.name)
    storage_client = storage.Client()

    object_name = "entregas_3.xlsx"
    bucket = storage_client.bucket(BUCKET_NAME_DOWNLOAD_REPORT)

    blob = storage.Blob(object_name, bucket)
    #blob.upload_from_file(archive, content_type='application/xlsx')
    blob.upload_from_string(tmp.read(), content_type='xlsx')

    #wb.save('test_1.xlsx')
    return