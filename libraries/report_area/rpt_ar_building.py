from openpyxl.workbook import Workbook
from openpyxl import load_workbook


def rpt_ar_building():
    wb= Workbook()
    ws = wb.active

    ws1 = wb.create_sheet('NewSheet')
    ws2 = wb.create_sheet('Another', 0)

    ws.title = 'MySheet'

    print(wb.sheetnames)

    return