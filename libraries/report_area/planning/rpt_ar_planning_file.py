
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, colors, Color, Alignment, PatternFill, GradientFill, Border, Side
from openpyxl.styles import NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

from tempfile import NamedTemporaryFile

from google.cloud import storage
from libraries.settings import BUCKET_NAME_DOWNLOAD_REPORT

def rpt_ar_planning_file(tmp_proyectos_planeacion, region, cut_date):

    data = {'tpp_hito':  ['IV', 'IP', 'IC', 'IE', 'DC', 'GASUE', 'GAS', 'SP', 'EL',  'AC'],
                'tpp_order_hito': [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
            }

    hitos = pd.DataFrame(data)

    tmp_proyectos_planeacion=pd.merge(tmp_proyectos_planeacion,hitos, on='tpp_hito', how="left",)
    tmp_proyectos_planeacion=tmp_proyectos_planeacion.sort_values(by=["tpp_proyecto","tpp_etapa", "tpp_order_hito"],ascending=True)

    tmp_proyectos_planeacion_excel=tmp_proyectos_planeacion.reindex(columns=[
                                                            'tpp_proyecto',
                                                            'tpp_hito',
                                                            'tpp_etapa',
                                                            'tpp_tarea_consume_buffer',
                                                            'tpp_avance_cc',
                                                            'tpp_avance_comparativo_semana',
                                                            'tpp_consumo_buffer',
                                                            'tpp_consumo_buffer_comparativo',
                                                            'tpp_fin_proyectado_optimista',
                                                            'tpp_fin_proyectado_pesimista',
                                                            'tpp_fin_programada',
                                                            'tpp_dias_atraso',
                                                            'tpp_ultima_semana',
                                                            'tpp_ultimo_mes',
                                                            'tpp_consumo_buffer_color'
                                                            ])
    
    
    
    wb=Workbook()
    ws=wb.active

    chart_title = NamedStyle(name= 'title')
    chart_title.font=Font(bold=True, color='0000FF', size=14)
    chart_title.alignment=Alignment(horizontal='left', vertical='center')

    chart_date =NamedStyle(name='chart_date')
    chart_date.font=Font(bold=True, color='000000', size=12)
    chart_date.alignment=Alignment(horizontal='left', vertical='center')

    column_title =NamedStyle(name='column_title')
    column_title.font=Font(bold=True, color='000000', size=10)
    column_title.alignment=Alignment(horizontal='center', vertical='center', wrap_text=True)
    column_title.fill= PatternFill('solid',fgColor='D3BC5F')
    bd=Side(style='thick', color='000000')
    column_title.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    table_body =NamedStyle(name='table_body')
    table_body.font=Font(color='000000', size=10)
    table_body.alignment=Alignment(horizontal='left', vertical='center', wrap_text=False)
    bd=Side(style='thin', color='000000')
    table_body.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    table_body_centered =NamedStyle(name='table_body_centered')
    table_body_centered.font=Font(color='000000', size=10)
    table_body_centered.alignment=Alignment(horizontal='center', vertical='center', wrap_text=True)
    bd=Side(style='thin', color='000000')
    table_body_centered.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    table_body_red=NamedStyle(name='table_body_red')
    table_body_red.font=Font(color='000000', size=10)
    table_body_red.alignment=Alignment(horizontal='left', vertical='center', wrap_text=False)
    table_body_red.fill= PatternFill('solid',fgColor='FF0000')
    bd=Side(style='thin', color='000000')
    table_body_red.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    table_body_yellow=NamedStyle(name='table_body_yellow')
    table_body_yellow.font=Font(color='000000', size=10)
    table_body_yellow.alignment=Alignment(horizontal='left', vertical='center', wrap_text=False)
    table_body_yellow.fill= PatternFill('solid',fgColor='FFFF00')
    bd=Side(style='thin', color='000000')
    table_body_yellow.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    table_body_green=NamedStyle(name='table_body_green')
    table_body_green.font=Font(color='000000', size=10)
    table_body_green.alignment=Alignment(horizontal='left', vertical='center', wrap_text=False)
    table_body_green.fill= PatternFill('solid',fgColor='008000')
    bd=Side(style='thin', color='000000')
    table_body_green.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    table_body_text_red=NamedStyle(name='table_body_text_red')
    table_body_text_red.font=Font(color='FF0000', size=10)
    table_body_text_red.alignment=Alignment(horizontal='left', vertical='center', wrap_text=False)
    bd=Side(style='thin', color='000000')
    table_body_text_red.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    table_body_text_yellow=NamedStyle(name='table_body_text_yellow')
    table_body_text_yellow.font=Font(color='FFFF00', size=10)
    table_body_text_yellow.alignment=Alignment(horizontal='left', vertical='center', wrap_text=False)
    bd=Side(style='thin', color='000000')
    table_body_text_yellow.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    table_body_text_green=NamedStyle(name='table_body_text_green')
    table_body_text_green.font=Font(color='008000', size=10)
    table_body_text_green.alignment=Alignment(horizontal='left', vertical='center', wrap_text=False)
    bd=Side(style='thin', color='000000')
    table_body_text_green.border = Border(left=bd, top=bd, right=bd, bottom=bd)





    
    ws['A1'].style = chart_title
    ws['A2'].style = chart_date
    ws['B2'].style = chart_date
    ws['A4'].style = column_title
    ws['B4'].style = column_title
    ws['C4'].style = column_title
    ws['D4'].style = column_title
    ws['E4'].style = column_title
    ws['F4'].style = column_title
    ws['G4'].style = column_title
    ws['H4'].style = column_title
    ws['I4'].style = column_title
    ws['J4'].style = column_title
    ws['K4'].style = column_title
    ws['L4'].style = column_title
    ws['M4'].style = column_title
    ws['N4'].style = column_title

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 9
    ws.column_dimensions['C'].width = 11
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 7
    ws.column_dimensions['F'].width = 11
    ws.column_dimensions['G'].width = 11
    ws.column_dimensions['H'].width = 11
    ws.column_dimensions['I'].width = 13
    ws.column_dimensions['J'].width = 13
    ws.column_dimensions['K'].width = 13
    ws.column_dimensions['L'].width = 7
    ws.column_dimensions['M'].width = 7
    ws.column_dimensions['N'].width = 7


    ws['A1']='Consolidado ' + region + ' Proyectos Gerencia de Planeación'
    ws['A2']='Fecha Corte'
    ws['B2']=cut_date.strftime('%d-%m-%Y')
    ws['A4']='Nombre Proyecto'
    ws['B4']='HITO'
    ws['C4']='Etapa'
    ws['D4']='Descripcion Tarea Consume Buffer'
    ws['E4']='% Avance CC'
    ws['F4']='% Avance CC Semana Anterior'
    ws['G4']='% Consumo Buffer'
    ws['H4']='% Consumo Buffer Semana Anterior'
    ws['I4']='Fecha Fin Proyectada Optimista'
    ws['J4']='Fecha Fin Proyectada Pesimista'
    ws['K4']='Fecha Fin Programada'
    ws['L4']='Dias de Atraso'
    ws['M4']='Ultima Semana'
    ws['N4']='Ultimo Mes'
    ws.merge_cells("A1:N1")
    ws.merge_cells("B2:C2")
    

    rows=dataframe_to_rows(tmp_proyectos_planeacion_excel, index=False)
    first_iteration=True
    iteration=1
    project_name=""
    project_group_row_start=4
    stage=""
    stage_group_row_start=4
    new_project=False

    for r_idx, row in enumerate(rows,4):
        for c_idx, col in enumerate(row, 1):
            if first_iteration!=True:
                if iteration == 1:
                    if project_name != col:

                        cells_to_merge="A"+str(project_group_row_start)+":A"+str(r_idx-1)
                        cell=ws.merge_cells(cells_to_merge)
                        new_project=True
                        project_name = col
                        project_group_row_start=r_idx
                        cell=ws.cell(row=r_idx, column=c_idx, value=col)
                        cell.style = table_body_centered
                elif iteration == 2:
                    cell=ws.cell(row=r_idx, column=c_idx, value=col)
                    cell.style = table_body_centered
                elif iteration == 3:
                    if stage != col or new_project==True:
                        #if  == 0:
                        #    cell=ws.cell(row=r_idx+ , column=5, value=0)
                        #    cell.style = table_body
                        #    cell=ws.cell(row=r_idx+ , column=7, value=0)
                        #    cell.style = table_body
                        #    cell=ws.cell(row=r_idx+ , column=13, value=0)
                        #    cell.style = table_body
                        #    cell=ws.cell(row=r_idx+ , column=14, value=0)
                        #    cell.style = table_body

                        cells_to_merge="C"+str(stage_group_row_start)+":C"+str(r_idx-1)

                        ws.merge_cells(cells_to_merge)
                        stage = col
                        stage_group_row_start=r_idx
                        cell=ws.cell(row=r_idx , column=c_idx, value=col)
                        cell.style = table_body_centered
                        new_project=False
                elif iteration == 4:
                    cell=ws.cell(row=r_idx, column=c_idx, value=col)
                    if col == "TERMINADO":
                        cell.style = table_body_text_green
                    else:
                        cell.style = table_body
                elif iteration == 5:
                    cell=ws.cell(row=r_idx , column=c_idx, value=round(col,2)*100)
                    cell.style = table_body
                elif iteration == 6 or iteration == 8:
                    if col==1:
                        cell=ws.cell(row=r_idx , column=c_idx, value="Sube")
                    elif col ==-1:
                        cell=ws.cell(row=r_idx , column=c_idx, value="Baja")
                    else:
                        cell=ws.cell(row=r_idx , column=c_idx, value="Igual")
                    cell.style = table_body
                elif iteration == 7:
                    cell=ws.cell(row=r_idx , column=c_idx, value=round(col,2)*100)
                        
                elif iteration == 9 or iteration == 10 or iteration == 11:
                    try:
                        info=col.strftime('%d-%m-%Y')
                    except:
                        info=""
                    cell=ws.cell(row=r_idx, column=c_idx, value=info)
                    cell.style = table_body
                elif iteration==12:
                    cell=ws.cell(row=r_idx, column=c_idx, value=col)
                    if col < 0:
                        cell.style = table_body_text_red
                    else:
                        cell.style = table_body_text_green
                elif iteration == 13:
                    cell=ws.cell(row=r_idx , column=c_idx, value=round(col,2)*100)
                    cell.style = table_body
                elif iteration == 14:
                    cell=ws.cell(row=r_idx , column=c_idx, value=round(col,2)*100)
                    cell.style = table_body
                elif iteration == 15:
                    cell=ws.cell(row=r_idx, column=7)
                    if col==1:
                        cell.style = table_body_green
                    elif col ==-1:
                        cell.style = table_body_red
                    else:
                        cell.style = table_body_yellow
                else:
                    cell=ws.cell(row=r_idx, column=c_idx, value=col)
                    cell.style = table_body
                iteration=iteration+1
            
        first_iteration=False
        iteration=1
    cells_to_merge="A"+str(project_group_row_start)+":A"+str(len(tmp_proyectos_planeacion_excel)+4)
    cell=ws.merge_cells(cells_to_merge)
    cells_to_merge="C"+str(stage_group_row_start)+":C"+str(len(tmp_proyectos_planeacion_excel)+4)
    
    ws.merge_cells(cells_to_merge)

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        storage_client = storage.Client()
    
        object_name = "corte_" +  cut_date.strftime('%d-%m-%Y') + "_planeacion_" + region + ".xlsx"
        bucket = storage_client.bucket(BUCKET_NAME_DOWNLOAD_REPORT)

        blob = storage.Blob(object_name, bucket)
        #blob.upload_from_file(archive, content_type='application/xlsx')
        blob.upload_from_string(tmp.read(), content_type='xlsx')

    #wb.save('test_1.xlsx')

    return