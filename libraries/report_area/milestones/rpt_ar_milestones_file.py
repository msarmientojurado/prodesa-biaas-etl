from openpyxl.drawing.image import Image
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.styles import NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from tempfile import NamedTemporaryFile
from libraries.settings import BUCKET_NAME_DOWNLOAD_REPORT

def rpt_ar_milestones_file(tbl_inicio_venta, 
            tbl_inicio_promesa, 
            tbl_inicio_construccion, 
            tbl_inicio_escrituracion,
            cut_date,
            region):
    
    tbl_inicio_venta_excel=tbl_inicio_venta.reindex(columns=[
                                                    'tiv_proyecto',
                                                    'tiv_etapa',
                                                    'tiv_dias_atraso',
                                                    'tiv_inicio_ventas_proyectado',
                                                    'tiv_inicio_ventas_programado',
                                                    'tiv_fiv_proyectado',
                                                    'tiv_fiv_programado',
                                                    'tiv_ppto_revisado_proyectado',
                                                    'tiv_ppto_revisado_programado',
                                                    'tiv_docs_ppto_proyectado',
                                                    'tiv_docs_ppto_programado',
                                                    'tiv_encargo_fiduciario_proyectado',
                                                    'tiv_encargo_fiduciario_programado',
                                                    'tiv_kit_comercial_proyectado',
                                                    'tiv_kit_comercial_programado',
                                                    'tiv_val_sv_model_proyectado',
                                                    'tiv_val_sv_model_programado',
                                                    'tiv_const_sv_model_proyectado',
                                                    'tiv_const_sv_model_programado',
                                                    'tiv_aprobac_lc_proyectado',
                                                    'tiv_aprobacion_lc_programado',
                                                    'tiv_radicacion_lc_proyectado',
                                                    'tiv_radicacion_lc_programado',
                                                    'tiv_salida_ventas_proyectado',
                                                    'tiv_salida_ventas_programado',
                                                    'tiv_acta_constituc_proyectado',
                                                    'tiv_acta_constituc_programado',
                                                    'tiv_visto_bueno_proyectado',
                                                    'tiv_visto_bueno_programado',
                                                    'tiv_elab_alternativ_proyectado',
                                                    'tiv_elab_alternativ_programado',
                                                    'tiv_prod_objetivo_proyectado',
                                                    'tiv_prod_objetivo_programado',
                                                    'tiv_lluvia_ideas_proyectado',
                                                    'tiv_lluvia_ideas_programado',
                                                    'tiv_kit_desarrollos_proyectado',
                                                    'tiv_kit_desarrollos_programado'])
    
    wb=Workbook()
    ws=wb.active

    chart_title = NamedStyle(name= 'title')
    chart_title.font=Font(bold=True, color='000080', size=24)
    chart_title.alignment=Alignment(horizontal='left', vertical='center')

    chart_subtitle = NamedStyle(name= 'char_subtitle')
    chart_subtitle.font=Font(bold=True, color='000000', size=22)
    chart_subtitle.alignment=Alignment(horizontal='left', vertical='center')

    chart_date =NamedStyle(name='chart_date')
    chart_date.font=Font(bold=True, color='000080', size=20)
    chart_date.alignment=Alignment(horizontal='left', vertical='center')

    column_title =NamedStyle(name='column_title')
    column_title.font=Font(bold=True, color='000000', size=14)
    column_title.alignment=Alignment(horizontal='center', vertical='center', wrap_text=True)
    column_title.fill= PatternFill('solid',fgColor='D3BC5F')
    bd=Side(style='thick', color='000000')
    column_title.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    table_body =NamedStyle(name='table_body')
    table_body.font=Font(color='000000', size=14)
    table_body.alignment=Alignment(horizontal='left', vertical='center', wrap_text=False)
    bd=Side(style='thin', color='000000')
    table_body.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    table_body_days_overdue =NamedStyle(name='table_body_days_overdue')
    table_body_days_overdue.font=Font(bold=True, color='000000', size=22)
    table_body_days_overdue.alignment=Alignment(horizontal='center', vertical='center', wrap_text=False)
    bd=Side(style='thin', color='000000')
    table_body_days_overdue.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    table_body_centered =NamedStyle(name='table_body_centered')
    table_body_centered.font=Font(color='000000', size=14)
    table_body_centered.alignment=Alignment(horizontal='center', vertical='center', wrap_text=True)
    bd=Side(style='thin', color='000000')
    table_body_centered.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    table_body_red=NamedStyle(name='table_body_red')
    table_body_red.font=Font(color='000000', size=14)
    table_body_red.alignment=Alignment(horizontal='center', vertical='center', wrap_text=False)
    table_body_red.fill= PatternFill('solid',fgColor='FF0000')
    bd=Side(style='thin', color='000000')
    table_body_red.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    table_body_yellow=NamedStyle(name='table_body_yellow')
    table_body_yellow.font=Font(color='000000', size=14)
    table_body_yellow.alignment=Alignment(horizontal='center', vertical='center', wrap_text=False)
    table_body_yellow.fill= PatternFill('solid',fgColor='FFFF00')
    bd=Side(style='thin', color='000000')
    table_body_yellow.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    table_body_green=NamedStyle(name='table_body_green')
    table_body_green.font=Font(color='000000', size=14)
    table_body_green.alignment=Alignment(horizontal='center', vertical='center', wrap_text=False)
    table_body_green.fill= PatternFill('solid',fgColor='008000')
    bd=Side(style='thin', color='000000')
    table_body_green.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    ws['A8'].style = chart_title
    ws['A9'].style = chart_date
    ws['B9'].style = chart_date
    ws['A11'].style = chart_subtitle
    ws['A12'].style = column_title
    ws['B12'].style = column_title
    ws['C12'].style = column_title
    ws['D12'].style = column_title
    ws['E12'].style = column_title
    ws['F12'].style = column_title
    ws['G12'].style = column_title
    ws['H12'].style = column_title
    ws['I12'].style = column_title
    ws['J12'].style = column_title
    ws['K12'].style = column_title
    ws['L12'].style = column_title
    ws['M12'].style = column_title
    ws['N12'].style = column_title
    ws['O12'].style = column_title
    ws['P12'].style = column_title
    ws['Q12'].style = column_title
    ws['R12'].style = column_title
    ws['S12'].style = column_title
    ws['T12'].style = column_title

    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12 
    ws.column_dimensions['D'].width = 19
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 22
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 18
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 18
    ws.column_dimensions['M'].width = 19
    ws.column_dimensions['N'].width = 17
    ws.column_dimensions['O'].width = 19
    ws.column_dimensions['P'].width = 20
    ws.column_dimensions['Q'].width = 18
    ws.column_dimensions['R'].width = 18
    ws.column_dimensions['S'].width = 18
    ws.column_dimensions['T'].width = 17



    ws['A8']='CUADRO DE HITOS DE PLANEACIÓN REGIONAL '+ region
    ws['A9']='FECHA'
    ws['B9']=cut_date.strftime('%d-%m-%Y')
    ws['A11']='INICIO VENTAS'
    ws['A12']='Nombre Proyecto'
    ws['B12']='Etapa'
    ws['C12']='# Dias de atraso'
    ws['D12']='Inicio de Ventas'
    ws['E12']='Factibilidad de Inicio de Ventas'
    ws['F12']='Ppto Definitivo (Tipo FIV)'
    ws['G12']='Entrega de Documentos para elaboración del presupuesto'
    ws['H12']='Encargo Fiduciario'
    ws['I12']='Entrega 1 a Comercial'
    ws['J12']='Validación SV y Modelos'
    ws['K12']='Construcción SV y Modelos'
    ws['L12']='Aprobación LC'
    ws['M12']='Radicación LC'
    ws['N12']='Diseño para salida a ventas'
    ws['O12']='Acta de Constitución'
    ws['P12']='VoBo (Visto Bueno)-Unidad estratégica y vicepresidente'
    ws['Q12']='Elaboración de Alternativas'
    ws['R12']='Definición de producto objetivo'
    ws['S12']='Lluvia de Ideas'
    ws['T12']='Kit de Desarrollos'

    ws.merge_cells("A8:S8")
    ws.merge_cells("B9:C9")
    ws.merge_cells("A11:E11")

    img_label=Image('libraries/report_area/milestones/img/rotulo_cuadro_hitos.JPG')
    ws.add_image(img_label, 'G1')

    rows=dataframe_to_rows(tbl_inicio_venta_excel, index=False)
    first_iteration=True
    iteration=1
    project_name=""
    project_group_row_start=12
    acumulator=0
    style_checked=False

    for r_idx, row in enumerate(rows,11):
        for c_idx, col in enumerate(row, 1):
            if first_iteration != True:
                if iteration == 1:
                    if project_name != col: 
                        cell=ws.cell(row=r_idx+acumulator, column=c_idx, value=col)
                        cell.style = table_body_centered
                        cells_to_merge="A"+str(project_group_row_start)+":A"+str(r_idx+acumulator-1)
                        ws.merge_cells(cells_to_merge)
                        project_group_row_start=r_idx+acumulator
                        project_name = col
                elif iteration == 2 or iteration == 3:
                    cell=ws.cell(row=r_idx+acumulator, column=c_idx, value=col)
                    if iteration == 2:
                        cells_to_merge="B"+str(r_idx+acumulator)+":B"+str(r_idx+acumulator +1)
                        cell.style = table_body_centered
                    elif iteration == 3:
                        cells_to_merge="C"+str(r_idx+acumulator)+":C"+str(r_idx+acumulator +1)
                        cell.style = table_body_days_overdue
                    ws.merge_cells(cells_to_merge)

                else:
                    info = date_format(col)
                    if iteration %2 ==0:
                        cell=ws.cell(row=r_idx+acumulator, column=c_idx-(iteration-4)/2, value=info)
                        if col < cut_date:
                            cell.style = table_body_green
                            style_checked = True
                        else:
                            cell.style = table_body_centered
                            style_checked = False
                    else:
                        cell=ws.cell(row=r_idx+acumulator+1, column=c_idx-1-(iteration-5)/2, value=info)
                        if style_checked == True:
                            cell.style = table_body_green
                        else:
                            if col < cut_date:
                                cell.style = table_body_red
                            elif col >= cut_date:
                                cell.style = table_body_yellow
                            else:
                                cell.style = table_body_centered
                    
            iteration=iteration+1
        iteration=1
        acumulator=acumulator + 1
        first_iteration = False

    cells_to_merge="A"+str(project_group_row_start)+":A"+str((len(tbl_inicio_venta_excel)*2)+12)
    #print(cells_to_merge)
    ws.merge_cells(cells_to_merge)

    #Inicio Promesa Section

    tbl_inicio_promesa_excel = tbl_inicio_promesa.reindex(columns=['tip_proyecto',
                                                        'tip_etapa',
                                                        'tip_dias_atraso',
                                                        'tip_inicio_promesas_proyectado',
                                                        'tip_inicio_promesas_programado',
                                                        'tip_ent_kit_prom_proyectado',
                                                        'tip_ent_kit_prom_programado',
                                                        'tip_permiso_ventas_proyectado',
                                                        'tip_permiso_ventas_programado',
                                                        'tip_min_hipo_reg_proyectado',
                                                        'tip_min_hipo_reg_programado',
                                                        'tip_rad_min_hipo_reg_proyectado',
                                                        'tip_rad_min_hipo_reg_programado',
                                                        'tip_cred_construct_proyectado',
                                                        'tip_cred_construct_programado',
                                                        'tip_linderos_proyectado',
                                                        'tip_linderos_programado',
                                                        'tip_fai_proyectado',
                                                        'tip_fai_programado',
                                                        'tip_constitut_urban_proyectado',
                                                        'tip_constitut_urban_programado'])
    row_promise_start= (len(tbl_inicio_venta)*2) + 12 + 2
    ws['A' + str(row_promise_start)].style = chart_subtitle
    ws['A' + str(row_promise_start + 1)].style = column_title
    ws['B' + str(row_promise_start + 1)].style = column_title
    ws['C' + str(row_promise_start + 1)].style = column_title
    ws['D' + str(row_promise_start + 1)].style = column_title
    ws['E' + str(row_promise_start + 1)].style = column_title
    ws['F' + str(row_promise_start + 1)].style = column_title
    ws['G' + str(row_promise_start + 1)].style = column_title
    ws['H' + str(row_promise_start + 1)].style = column_title
    ws['I' + str(row_promise_start + 1)].style = column_title
    ws['J' + str(row_promise_start + 1)].style = column_title
    ws['K' + str(row_promise_start + 1)].style = column_title
    ws['L' + str(row_promise_start + 1)].style = column_title

    ws['A' + str(row_promise_start)]='INICIO DE PROMESAS'
    ws['A' + str(row_promise_start + 1)]='Nombre del Proyecto'
    ws['B' + str(row_promise_start + 1)]='Etapa'
    ws['C' + str(row_promise_start + 1)]='# Dias de atraso'
    ws['D' + str(row_promise_start + 1)]='Inicio de Promesas'
    ws['E' + str(row_promise_start + 1)]='Entrega Kit Promesas'
    ws['F' + str(row_promise_start + 1)]='Permiso de Ventas'
    ws['G' + str(row_promise_start + 1)]='Minuta de Hipoteca Registrada'
    ws['H' + str(row_promise_start + 1)]='Radicación minuta Hipoteca a Registro'
    ws['I' + str(row_promise_start + 1)]='Credito Constructor'
    ws['J' + str(row_promise_start + 1)]='Linderos'
    ws['K' + str(row_promise_start + 1)]='FAI'
    ws['L' + str(row_promise_start + 1)]='Constitución de la Urbanización'
    ws.merge_cells("A"+str(row_promise_start)+":E"+str(row_promise_start))
    
    rows=dataframe_to_rows(tbl_inicio_promesa_excel, index=False)
    first_iteration=True
    iteration=1
    project_name=""
    project_group_row_start=row_promise_start + 1
    acumulator=0
    style_checked=False

    for r_idx, row in enumerate(rows,row_promise_start):
        for c_idx, col in enumerate(row, 1):
            if first_iteration != True:
                if iteration == 1:
                    if project_name != col: 
                        cell=ws.cell(row=r_idx+acumulator, column=c_idx, value=col)
                        cell.style = table_body_centered
                        cells_to_merge="A"+str(project_group_row_start)+":A"+str(r_idx+acumulator-1)
                        ws.merge_cells(cells_to_merge)
                        project_group_row_start=r_idx+acumulator
                        project_name = col
                elif iteration == 2 or iteration == 3:
                    cell=ws.cell(row=r_idx+acumulator, column=c_idx, value=col)
                    if iteration == 2:
                        cells_to_merge="B"+str(r_idx+acumulator)+":B"+str(r_idx+acumulator +1)
                        cell.style = table_body_centered
                    elif iteration == 3:
                        cells_to_merge="C"+str(r_idx+acumulator)+":C"+str(r_idx+acumulator +1)
                        cell.style = table_body_days_overdue
                    ws.merge_cells(cells_to_merge)

                else:
                    info = date_format(col)
                    if iteration %2 ==0:
                        cell=ws.cell(row=r_idx+acumulator, column=c_idx-(iteration-4)/2, value=info)
                        if col < cut_date:
                            cell.style = table_body_green
                            style_checked = True
                        else:
                            cell.style = table_body_centered
                            style_checked = False
                    else:
                        cell=ws.cell(row=r_idx+acumulator+1, column=c_idx-1-(iteration-5)/2, value=info)
                        if style_checked == True:
                            cell.style = table_body_green
                        else:
                            if col < cut_date:
                                cell.style = table_body_red
                            elif col >= cut_date:
                                cell.style = table_body_yellow
                            else:
                                cell.style = table_body_centered
                    
            iteration=iteration+1
        iteration=1
        acumulator=acumulator + 1
        first_iteration = False
    cells_to_merge="A"+str(project_group_row_start)+":A"+str(row_promise_start + (len(tbl_inicio_promesa_excel)*2)+1)
    #print(cells_to_merge)
    ws.merge_cells(cells_to_merge)

    #Inicio Construccion Section

    tbl_inicio_construccion_excel = tbl_inicio_construccion.reindex(columns=['tic_proyecto',
                                                                        'tic_etapa',
                                                                        'tic_dias_atraso',
                                                                        'tic_inicio_construccion_proyectado',
                                                                        'tic_inicio_construccion_programado',
                                                                        'tic_proc_contratacion_proyectado',
                                                                        'tic_proc_contratacion_programado',
                                                                        'tic_entrega_kit2_proyectado',
                                                                        'tic_entrega_kit2_programado',
                                                                        'tic_entrega_kit1_proyectado',
                                                                        'tic_entrega_kit1_programado',
                                                                        'tic_factib_inic_constru_proyectado',
                                                                        'tic_factib_inic_constru_programado',
                                                                        'tic_ppto_definitivo_proyectado',
                                                                        'tic_ppto_definitivo_programado',
                                                                        'tic_ppto_sipro_proyectado',
                                                                        'tic_ppto_sipro_programado',
                                                                        'tic_docs_inic_constru_proyectado',
                                                                        'tic_docs_inic_constru_programado',
                                                                        'tic_diseno_inic_constru_proyectado',
                                                                        'tic_diseno_inic_constru_programado'])
    row_building_start= row_promise_start + (len(tbl_inicio_promesa_excel)*2) + 3
    ws['A' + str(row_building_start)].style = chart_subtitle
    ws['A' + str(row_building_start + 1)].style = column_title
    ws['B' + str(row_building_start + 1)].style = column_title
    ws['C' + str(row_building_start + 1)].style = column_title
    ws['D' + str(row_building_start + 1)].style = column_title
    ws['E' + str(row_building_start + 1)].style = column_title
    ws['F' + str(row_building_start + 1)].style = column_title
    ws['G' + str(row_building_start + 1)].style = column_title
    ws['H' + str(row_building_start + 1)].style = column_title
    ws['I' + str(row_building_start + 1)].style = column_title
    ws['J' + str(row_building_start + 1)].style = column_title
    ws['K' + str(row_building_start + 1)].style = column_title
    ws['L' + str(row_building_start + 1)].style = column_title

    ws['A' + str(row_building_start)]='INICIO DE CONSTRUCCION'
    ws['A' + str(row_building_start + 1)]='Nombre Proyecto'
    ws['B' + str(row_building_start + 1)]='Etapa'
    ws['C' + str(row_building_start + 1)]='# Dias de atraso'
    ws['D' + str(row_building_start + 1)]='Inicio de Construcción'
    ws['E' + str(row_building_start + 1)]='Proceso Contratación'
    ws['F' + str(row_building_start + 1)]='Entrega Kit 2'
    ws['G' + str(row_building_start + 1)]='Entrega Kit 1'
    ws['H' + str(row_building_start + 1)]='Factibilidad de inicio de construcción'
    ws['I' + str(row_building_start + 1)]='Ppto Definitivo (Tipo FIC)'
    ws['J' + str(row_building_start + 1)]='Presupuesto Sipro'
    ws['K' + str(row_building_start + 1)]='Entrega de documentos para elaboración de presupuesto de inicio de construcción'
    ws['L' + str(row_building_start + 1)]='Diseño para inicio de construcción'
    
    ws.merge_cells("A"+str(row_building_start)+":E"+str(row_building_start))

    rows=dataframe_to_rows(tbl_inicio_construccion_excel , index=False)
    first_iteration=True
    iteration=1
    project_name=""
    project_group_row_start=row_building_start + 1
    acumulator=0
    style_checked=False

    for r_idx, row in enumerate(rows,row_building_start):
        for c_idx, col in enumerate(row, 1):
            if first_iteration != True:
                if iteration == 1:
                    if project_name != col: 
                        cell=ws.cell(row=r_idx+acumulator, column=c_idx, value=col)
                        cell.style = table_body_centered
                        cells_to_merge="A"+str(project_group_row_start)+":A"+str(r_idx+acumulator-1)
                        ws.merge_cells(cells_to_merge)
                        project_group_row_start=r_idx+acumulator
                        project_name = col
                elif iteration == 2 or iteration == 3:
                    cell=ws.cell(row=r_idx+acumulator, column=c_idx, value=col)
                    if iteration == 2:
                        cells_to_merge="B"+str(r_idx+acumulator)+":B"+str(r_idx+acumulator +1)
                        cell.style = table_body_centered
                    elif iteration == 3:
                        cells_to_merge="C"+str(r_idx+acumulator)+":C"+str(r_idx+acumulator +1)
                        cell.style = table_body_days_overdue
                    ws.merge_cells(cells_to_merge)

                else:
                    info = date_format(col)
                    if iteration %2 ==0:
                        cell=ws.cell(row=r_idx+acumulator, column=c_idx-(iteration-4)/2, value=info)
                        if col < cut_date:
                            cell.style = table_body_green
                            style_checked = True
                        else:
                            cell.style = table_body_centered
                            style_checked = False
                    else:
                        cell=ws.cell(row=r_idx+acumulator+1, column=c_idx-1-(iteration-5)/2, value=info)
                        if style_checked == True:
                            cell.style = table_body_green
                        else:
                            if col < cut_date:
                                cell.style = table_body_red
                            elif col >= cut_date:
                                cell.style = table_body_yellow
                            else:
                                cell.style = table_body_centered
                    
            iteration=iteration+1
        iteration=1
        acumulator=acumulator + 1
        first_iteration = False
    cells_to_merge="A"+str(project_group_row_start)+":A"+str(row_building_start + (len(tbl_inicio_construccion_excel)*2)+1)
    #print(cells_to_merge)
    ws.merge_cells(cells_to_merge)

    #Inicio Escrituración Section

    tbl_inicio_escrituracion_excel = tbl_inicio_escrituracion.reindex(columns=['tie_proyecto',
                                                                        'tie_etapa',
                                                                        'tie_dias_atraso',
                                                                        'tie_inicio_escrituracion_proyectado',
                                                                        'tie_inicio_escrituracion_programado',
                                                                        'tie_poder_fiduciaria_proyectado',
                                                                        'tie_poder_fiduciaria_programado',
                                                                        'tie_salida_rph_proyectado',
                                                                        'tie_salida_rph_programado',
                                                                        'tie_cierre_rph_proyectado',
                                                                        'tie_cierre_rph_programado',
                                                                        'tie_licencia_ph_proyectado',
                                                                        'tie_licencia_ph_programado',
                                                                        'tie_modificacion_lc_proyectado',
                                                                        'tie_modificacion_lc_programado',
                                                                        'tie_radic_modif_lc_proyectado',
                                                                        'tie_radic_modif_lc_programado'])
    row_registration_start= row_building_start + (len(tbl_inicio_construccion_excel)*2) + 3
    ws['A' + str(row_registration_start)].style = chart_subtitle
    ws['A' + str(row_registration_start + 1)].style = column_title
    ws['B' + str(row_registration_start + 1)].style = column_title
    ws['C' + str(row_registration_start + 1)].style = column_title
    ws['D' + str(row_registration_start + 1)].style = column_title
    ws['E' + str(row_registration_start + 1)].style = column_title
    ws['F' + str(row_registration_start + 1)].style = column_title
    ws['G' + str(row_registration_start + 1)].style = column_title
    ws['H' + str(row_registration_start + 1)].style = column_title
    ws['I' + str(row_registration_start + 1)].style = column_title
    ws['J' + str(row_registration_start + 1)].style = column_title

    ws['A' + str(row_registration_start)]='INICIO DE ESCRITURACION'
    ws['A' + str(row_registration_start + 1)]='Nombre del Proyecto'
    ws['B' + str(row_registration_start + 1)]='Etapa'
    ws['C' + str(row_registration_start + 1)]='# Dias de atraso'
    ws['D' + str(row_registration_start + 1)]='Inicio Escrituración'
    ws['E' + str(row_registration_start + 1)]='Poder de la Fiduciaria'
    ws['F' + str(row_registration_start + 1)]='Salida RPH'
    ws['G' + str(row_registration_start + 1)]='Cerrar RPH'
    ws['H' + str(row_registration_start + 1)]='LICENCIAPH'
    ws['I' + str(row_registration_start + 1)]='Modificación LC'
    ws['J' + str(row_registration_start + 1)]='Radicación Modificación Licencia de Construcción'

    ws.merge_cells("A"+str(row_registration_start)+":E"+str(row_registration_start))

    
    rows=dataframe_to_rows(tbl_inicio_escrituracion_excel , index=False)
    first_iteration=True
    iteration=1
    project_name=""
    project_group_row_start=row_registration_start + 1
    acumulator=0
    style_checked=False

    for r_idx, row in enumerate(rows,row_registration_start):
        for c_idx, col in enumerate(row, 1):
            if first_iteration != True:
                if iteration == 1:
                    if project_name != col: 
                        cell=ws.cell(row=r_idx+acumulator, column=c_idx, value=col)
                        cell.style = table_body_centered
                        cells_to_merge="A"+str(project_group_row_start)+":A"+str(r_idx+acumulator-1)
                        ws.merge_cells(cells_to_merge)
                        project_group_row_start=r_idx+acumulator
                        project_name = col
                elif iteration == 2 or iteration == 3:
                    cell=ws.cell(row=r_idx+acumulator, column=c_idx, value=col)
                    if iteration == 2:
                        cells_to_merge="B"+str(r_idx+acumulator)+":B"+str(r_idx+acumulator +1)
                        cell.style = table_body_centered
                    elif iteration == 3:
                        cells_to_merge="C"+str(r_idx+acumulator)+":C"+str(r_idx+acumulator +1)
                        cell.style = table_body_days_overdue
                    ws.merge_cells(cells_to_merge)

                else:
                    info = date_format(col)
                    if iteration %2 ==0:
                        cell=ws.cell(row=r_idx+acumulator, column=c_idx-(iteration-4)/2, value=info)
                        if col < cut_date:
                            cell.style = table_body_green
                            style_checked = True
                        else:
                            cell.style = table_body_centered
                            style_checked = False
                    else:
                        cell=ws.cell(row=r_idx+acumulator+1, column=c_idx-1-(iteration-5)/2, value=info)
                        if style_checked == True:
                            cell.style = table_body_green
                        else:
                            if col < cut_date:
                                cell.style = table_body_red
                            elif col >= cut_date:
                                cell.style = table_body_yellow
                            else:
                                cell.style = table_body_centered
                    
            iteration=iteration+1
        iteration=1
        acumulator=acumulator + 1
        first_iteration = False
    cells_to_merge="A"+str(project_group_row_start)+":A"+str(row_registration_start + (len(tbl_inicio_escrituracion_excel)*2)+1)
    #print(cells_to_merge)
    ws.merge_cells(cells_to_merge)

    tmp = NamedTemporaryFile()
    wb.save(tmp.name)
    return tmp

def date_format(date_to_format):
    info=""
    try:
        info=date_to_format.strftime('%d-%m-%Y')
        month_mapping={
            '01':'ene',
            '02':'feb',
            '03':'mar',
            '04':'abr',
            '05':'may',
            '06':'jun',
            '07':'jul',
            '08':'ago',
            '09':'sep',
            '10':'oct',
            '11':'nov',
            '12':'dic'
        }
        auxCol=info.split("-")
        month=month_mapping[auxCol[1]]
        info = auxCol[0]+"-"+month+"-"+auxCol[2]
    except:
        info=""
    return info