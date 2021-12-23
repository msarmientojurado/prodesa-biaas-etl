from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.styles import NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from tempfile import NamedTemporaryFile
from libraries.settings import BUCKET_NAME_DOWNLOAD_REPORT

def rpt_ar_commercial_file(tmp_proyectos_comercial, region, cut_date):
    
    tmp_proyectos_comercial_excel=tmp_proyectos_comercial.reindex(columns=[
                                                            'tpcm_proyecto',
                                                            'tpcm_etapa',
                                                            'tpcm_programacion',
                                                            'tpcm_tarea_consume_buffer',
                                                            'tpcm_avance_cc',
                                                            'tpcm_avance_comparativo_semana',
                                                            'tpcm_consumo_buffer',
                                                            'tpcm_consumo_buffer_comparativo',
                                                            'tpcm_fin_proyectado_optimista',
                                                            'tpcm_fin_proyectado_pesimista',
                                                            'tpcm_fin_programada',
                                                            'tpcm_dias_atraso',
                                                            'tpcm_ultima_semana',
                                                            'tpcm_ultimo_mes',
                                                            'tpcm_consumo_buffer_color'
                                                            ])
    
    wb=Workbook()
    ws=wb.active

    chart_title = NamedStyle(name= 'title')
    chart_title.font=Font(bold=True, color='000080', size=14)
    chart_title.alignment=Alignment(horizontal='left', vertical='center')

    chart_date =NamedStyle(name='chart_date')
    chart_date.font=Font(bold=True, color='000000', size=12)
    chart_date.alignment=Alignment(horizontal='left', vertical='center')

    column_title =NamedStyle(name='column_title')
    column_title.font=Font(bold=True, color='000000', size=10)
    column_title.alignment=Alignment(horizontal='center', vertical='center', wrap_text=True)
    column_title.fill= PatternFill('solid',fgColor='D3BC5F')
    bd=Side(style='thick', color='000000')
    column_title.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    table_body =NamedStyle(name='table_body')
    table_body.font=Font(color='000000', size=12)
    table_body.alignment=Alignment(horizontal='left', vertical='center', wrap_text=False)
    bd=Side(style='thin', color='000000')
    table_body.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    table_body_project_name =NamedStyle(name='table_body_project_name')
    table_body_project_name.font=Font(color='000000', size=10)
    table_body_project_name.alignment=Alignment(horizontal='center', vertical='center', wrap_text=True)
    bd=Side(style='thin', color='000000')
    table_body_project_name.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    table_body_centered =NamedStyle(name='table_body_centered')
    table_body_centered.font=Font(color='000000', size=12)
    table_body_centered.alignment=Alignment(horizontal='center', vertical='center', wrap_text=True)
    bd=Side(style='thin', color='000000')
    table_body_centered.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    table_body_red=NamedStyle(name='table_body_red')
    table_body_red.font=Font(color='000000', size=12)
    table_body_red.alignment=Alignment(horizontal='center', vertical='center', wrap_text=False)
    table_body_red.fill= PatternFill('solid',fgColor='FF0000')
    bd=Side(style='thin', color='000000')
    table_body_red.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    table_body_yellow=NamedStyle(name='table_body_yellow')
    table_body_yellow.font=Font(color='000000', size=12)
    table_body_yellow.alignment=Alignment(horizontal='center', vertical='center', wrap_text=False)
    table_body_yellow.fill= PatternFill('solid',fgColor='FFFF00')
    bd=Side(style='thin', color='000000')
    table_body_yellow.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    table_body_green=NamedStyle(name='table_body_green')
    table_body_green.font=Font(color='000000', size=12)
    table_body_green.alignment=Alignment(horizontal='center', vertical='center', wrap_text=False)
    table_body_green.fill= PatternFill('solid',fgColor='008000')
    bd=Side(style='thin', color='000000')
    table_body_green.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    table_body_text_red=NamedStyle(name='table_body_text_red')
    table_body_text_red.font=Font(color='FF0000', size=12)
    table_body_text_red.alignment=Alignment(horizontal='center', vertical='center', wrap_text=False)
    bd=Side(style='thin', color='000000')
    table_body_text_red.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    table_body_text_green=NamedStyle(name='table_body_text_green')
    table_body_text_green.font=Font(color='008000', size=12)
    table_body_text_green.alignment=Alignment(horizontal='center', vertical='center', wrap_text=False)
    bd=Side(style='thin', color='000000')
    table_body_text_green.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    bd_thick=Side(style='thick', color='000000')
    bd_thin=Side(style='thin', color='000000')
    bd_none=Side(style='none', color= '000000')

    ws['A1'].style = chart_title
    ws['A2'].style = chart_date
    ws['B2'].style = chart_date
    ws['A4'].style = column_title
    ws['B4'].style = column_title
    ws['C4'].style = column_title
    ws['D4'].style = column_title
    ws['E4'].style = column_title
    ws['F4'].style = column_title
    ws['G4'].style = column_title
    ws['H4'].style = column_title
    ws['I4'].style = column_title
    ws['J4'].style = column_title
    ws['K4'].style = column_title
    ws['L4'].style = column_title
    ws['M4'].style = column_title
    ws['N4'].style = column_title

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 13
    ws.column_dimensions['C'].width = 21
    ws.column_dimensions['D'].width = 33
    ws.column_dimensions['E'].width = 7
    ws.column_dimensions['F'].width = 11
    ws.column_dimensions['G'].width = 9
    ws.column_dimensions['H'].width = 11
    ws.column_dimensions['I'].width = 13
    ws.column_dimensions['J'].width = 13
    ws.column_dimensions['K'].width = 13
    ws.column_dimensions['L'].width = 8
    ws.column_dimensions['M'].width = 8
    ws.column_dimensions['N'].width = 8

    ws['A1']='Consolidado ' + region + ' Proyectos Gerencia Comercial'
    ws['A2']='Fecha Corte'
    ws['B2']=cut_date.strftime('%d-%m-%Y')
    ws['A4']='Nombre Proyecto'
    ws['B4']='Etapa'
    ws['C4']='Programación'
    ws['D4']='Descripción Tarea Consume Buffer'
    ws['E4']='% Avance CC'
    ws['F4']='% Avance CC Semana Anterior'
    ws['G4']='% Consumo Buffer'
    ws['H4']='% Consumo Buffer Semana Anterior'
    ws['I4']='Fecha Fin Proyectada Optimista'
    ws['J4']='Fecha Fin Proyectada Pesimista'
    ws['K4']='Fecha Fin Programada'
    ws['L4']='Días de Atraso'
    ws['M4']='Última Semana'
    ws['N4']='Último Mes'
    ws.merge_cells("A1:N1")
    ws.merge_cells("B2:C2")

    rows=dataframe_to_rows(tmp_proyectos_comercial_excel, index=False)
    first_iteration=True
    iteration=1
    project_name=""
    project_group_row_start=4
    stage=""
    stage_group_row_start=4
    new_project=False

    for r_idx, row in enumerate(rows,4):
        for c_idx, col in enumerate(row, 1):
            if first_iteration!=True:
                if iteration == 1:
                    if project_name != col:
                        cells_to_merge="A"+str(project_group_row_start)+":A"+str(r_idx-1)
                        cell=ws.cell(row=project_group_row_start, column=1)
                        cell.border = Border(left=bd_thick, top=bd_thick, right=bd_thin, bottom=bd_thick)
                        for bold_column in range(2,14):
                            cell=ws.cell(row=project_group_row_start, column=bold_column)
                            cell.border = Border(left=bd_thin, top=bd_thick, right=bd_thin, bottom=bd_thin)
                        for bold_row in range(project_group_row_start,r_idx + 1):
                            cell=ws.cell(row=bold_row, column=14)
                            if bold_row==project_group_row_start :
                                cell.border = Border(left=bd_thin, top=bd_thick, right=bd_thick, bottom=bd_thin)
                            elif bold_row == r_idx:
                                cell.border = Border(left=bd_thin, top=bd_thin, right=bd_thick, bottom = bd_thick)
                            else:
                                cell.border = Border(left=bd_thin, top=bd_thin, right=bd_thick, bottom=bd_thin)
                            
                        cell=ws.merge_cells(cells_to_merge)
                        new_project=True
                        project_name = col
                        project_group_row_start=r_idx
                        cell=ws.cell(row=r_idx, column=c_idx, value=col)
                        cell.style = table_body_project_name
                elif iteration == 2:
                    if stage != col or new_project==True:
                        cells_to_merge="B"+str(stage_group_row_start)+":B"+str(r_idx-1)
                        ws.merge_cells(cells_to_merge)
                        stage = col
                        stage_group_row_start=r_idx
                        cell=ws.cell(row=r_idx, column=c_idx, value=col)
                        cell.style = table_body_centered
                        new_project=False
                elif iteration == 4:
                    cell=ws.cell(row=r_idx, column=c_idx, value=col)
                    if col == "TERMINADO":
                        cell.style = table_body_text_green
                    else:
                        cell.style = table_body
                elif iteration == 5:
                    cell=ws.cell(row=r_idx, column=c_idx, value=round(col,2)*100)
                    cell.style = table_body_centered
                elif iteration == 6 or iteration == 8:
                    if col==1:
                        cell=ws.cell(row=r_idx, column=c_idx, value="Sube")
                    elif col ==-1:
                        cell=ws.cell(row=r_idx, column=c_idx, value="Baja")
                    else:
                        cell=ws.cell(row=r_idx, column=c_idx, value="Igual")
                    cell.style = table_body
                elif iteration == 7:
                    cell=ws.cell(row=r_idx, column=c_idx, value=round(col,2)*100)
                        
                elif iteration == 9 or iteration == 10 or iteration == 11:
                    info = date_format(col)
                    cell=ws.cell(row=r_idx, column=c_idx, value=info)
                    cell.style = table_body
                elif iteration==12:
                    cell=ws.cell(row=r_idx, column=c_idx, value=col)
                    if col < 0:
                        cell.style = table_body_text_red
                    else:
                        cell.style = table_body_text_green
                elif iteration == 13:
                    cell=ws.cell(row=r_idx, column=c_idx, value=round(col,2)*100)
                    cell.style = table_body_centered
                elif iteration == 14:
                    cell=ws.cell(row=r_idx, column=c_idx, value=round(col,2)*100)
                    cell.style = table_body_centered
                elif iteration == 15:
                    cell=ws.cell(row=r_idx, column=7)
                    if col==1:
                        cell.style = table_body_green
                    elif col ==-1:
                        cell.style = table_body_red
                    else:
                        cell.style = table_body_yellow
                else:
                    cell=ws.cell(row=r_idx, column=c_idx, value=col)
                    cell.style = table_body
                iteration=iteration+1
            
        first_iteration=False
        iteration=1
    cell=ws.cell(row=project_group_row_start, column=1)
    cell.border = Border(left=bd_thick, top=bd_thick, right=bd_thin, bottom=bd_thick)
    for bold_column in range(2,14):
        cell=ws.cell(row=project_group_row_start, column=bold_column)
        cell.border = Border(left=bd_thin, top=bd_thick, right=bd_thin, bottom=bd_thin)
    for bold_row in range(project_group_row_start,len(tmp_proyectos_comercial_excel) + 5):
        cell=ws.cell(row=bold_row, column=14)
        if bold_row==project_group_row_start :
            cell.border = Border(left=bd_thin, top=bd_thick, right=bd_thick, bottom=bd_thin)
        elif bold_row == r_idx:
            cell.border = Border(left=bd_thin, top=bd_thin, right=bd_thick, bottom = bd_thin)
        else:
            cell.border = Border(left=bd_thin, top=bd_thin, right=bd_thick, bottom=bd_thin)
    cells_to_merge="A"+str(project_group_row_start)+":A"+str(len(tmp_proyectos_comercial_excel)+4)
    cell=ws.merge_cells(cells_to_merge)
    cells_to_merge="B"+str(stage_group_row_start)+":B"+str(len(tmp_proyectos_comercial_excel)+4)
    ws.merge_cells(cells_to_merge)

    for bold_column in range(2,15):
        cell=ws.cell(row=len(tmp_proyectos_comercial_excel)+5, column=bold_column)
        if bold_column < 14:
            cell.border = Border(left=bd_none, top=bd_thick, right=bd_none, bottom = bd_none)
        else:
            cell.border = Border(left=bd_none, top=bd_thick, right=bd_none, bottom = bd_none)

    tmp = NamedTemporaryFile()
    wb.save(tmp.name)

    return tmp

def date_format(date_to_format):
    info=""
    try:
        info=date_to_format.strftime('%d-%m-%Y')
        month_mapping={
            '01':'ene',
            '02':'feb',
            '03':'mar',
            '04':'abr',
            '05':'may',
            '06':'jun',
            '07':'jul',
            '08':'ago',
            '09':'sep',
            '10':'oct',
            '11':'nov',
            '12':'dic'
        }
        auxCol=info.split("-")
        month=month_mapping[auxCol[1]]
        info = auxCol[0]+"-"+month+"-"+auxCol[2]
    except:
        info=""
    return info